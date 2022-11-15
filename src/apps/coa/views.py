import datetime
import os
import traceback
import urllib

from openpyxl import load_workbook

from selenium.common.exceptions import TimeoutException

from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.contrib.auth import login
from django.conf import settings

from apps.log.models import TracebackLog
from apps.user.models import CustomUser, DatabaseControl, CustomSetting, AnyToken

from apps.coa.apis import *
from apps.coa.builders import CropPriceOriginBuilder, CropProduceTotalBuilder
from apps.coa.utils import CustomError, get_driver
from apps.coa.models import ProductCode, CropProduceUnit, LivestockByproduct, CropProduceTotal


def api_view(command_text):
    # 將輸入的文字依據空格分割
    list_text = command_text.strip().split(' ')
    list_params = list()
    for text in list_text:
        if len(text) > 0:
            list_params.append(text)

    command = list_params[0]
    # 判斷分割後的第一組文字是哪一個指令
    if command in ["產值", "總產值"]:
        apiview = ProduceValueApiView
    elif command in ["毛額", "生產毛額"]:
        apiview = GrossApiView
    elif command in ["所得", "農家所得"]:
        apiview = IncomeApiView
    elif command  in ["農牧戶人口數", "農牧戶戶數", "人口數", "戶數", "農牧戶"]:
        apiview = FarmerApiView
    elif command in ["耕地面積"]:
        apiview = FarmerAreaApiView
    elif command in ["勞動力", "就業人口"]:
        apiview = LaborforceApiView
    elif command in ["災害"]:
        apiview = DisasterApiView
    elif command in ["農保", "津貼", "老農津貼", "獎助學金"]:
        apiview = WelfareApiView
    elif command in ["產地", "批發"]:
        apiview = CropPriceApiView
    elif command in ["種植面積", "單位產值", "單位產量"]:
        apiview = CropProduceApiView
    elif command in ["成本", "生產成本", "費用", "生產費用", "粗收益", "淨收入率", "工時"]:
        apiview = CropCostApiView
    elif command in ["毛豬", "交易量", "價格", "重量"]:
        apiview = LivestockHogApiView
    elif command in ["拍賣價", "產地價", "零售價"]:
        apiview = LivestockPriceApiView
    elif command in ["場數", "飼養場數"]:
        apiview = LivestockFeedlotApiView
    elif command in ["在養", "在養量"]:
        apiview = LivestockFeedamountApiView
    elif command in ["屠宰", "屠宰量"]:
        apiview = LivestockSlaughterApiView
    elif command in ["副產", "副產品", "副產物"]:
        apiview = LivestockByproductApiView
    elif command in ["代碼", "作物代碼"]:
        apiview = ProductCodeApiView
    elif command in ["產量"]:
        reply = f"無法搜尋「{command_text}」\n\n"
        reply += f"產量（作物產量）的指令為「產量 縣市 品項 年份」可加上鄉鎮「產量 縣市鄉鎮 品項 年份」，例如：\n"
        reply += f"「產量 雲林 落花生 108」\n「產量 雲林土庫 落花生 108」\n\n"
        reply += f"產量（副產物產量）的指令為「產量 品項 年份」或「產量 品項 年份 縣市」，也可使用替代指令「副產物」或「副產品」，例如：\n"
        reply += f"「產量 雞蛋 108」\n「產量 牛乳 109 雲林」\n"
        reply += f"「副產物 蜂蜜 108」\n「副產品 蜂蜜 105 彰化」"
        if not 3 <= len(list_params) <= 4:
            return reply

        apiview = None
        if len(list_params) == 3:
            command_text = command_text.replace("產量", "產量（副產物產量）")
            apiview = LivestockByproductApiView

        if len(list_params) == 4:
            product = list_params[1]
            query_set = LivestockByproduct.objects.filter(name__icontains=product, sub_class="product")
            if query_set.count() > 0:
                command_text = command_text.replace("產量", "產量（副產物產量）")
                apiview = LivestockByproductApiView

            query_set = CropProduceTotal.objects.filter(name__icontains=product)
            if query_set.count() > 0:
                command_text = command_text.replace("產量", "產量（作物產量）")
                apiview = CropProduceApiView

            product = list_params[2]
            query_set = LivestockByproduct.objects.filter(name__icontains=product, sub_class="product")
            if query_set.count() > 0:
                command_text = command_text.replace("產量", "產量（副產物產量）")
                apiview = LivestockByproductApiView

            query_set = CropProduceTotal.objects.filter(name__icontains=product)
            if query_set.count() > 0:
                command_text = command_text.replace("產量", "產量（作物產量）")
                apiview = CropProduceApiView

            if apiview is None:
                reply = f"無法搜尋「{command_text}」\n" + f"查無品項，請修改品項關鍵字後重新查詢"
                return reply

    elif "指令" in command_text:
        reply = f"直接輸入指令可以查詢使用方式(括號內為備註不需輸入)，目前提供的指令如下\n\n"
        reply += "經濟指標及其他類：\n"
        reply += "產值\n"
        reply += "總產值\n"
        reply += "生產毛額\n"
        reply += "農家所得\n"
        reply += "農牧戶(包含戶數及人口數)\n"
        reply += "耕地面積\n"
        reply += "就業人口\n"
        reply += "災害\n"
        reply += "農保\n"
        reply += "老農津貼\n"
        reply += "獎助學金\n\n"

        reply += "農耕類：\n"
        reply += "產地\n"
        reply += "批發\n"
        reply += "產量(作物產量)\n"
        reply += "種植面積\n"
        reply += "單位產值\n"
        reply += "單位產量\n"
        reply += "生產成本(包含生產費用、粗收益、淨收入率、工時)\n\n"

        reply += "畜禽類：\n"
        reply += "毛豬(包含交易量、拍賣價、平均重量)\n"
        reply += "拍賣價\n"
        reply += "產地價\n"
        reply += "零售價\n"
        reply += "飼養場數\n"
        reply += "在養量\n"
        reply += "屠宰量\n"
        reply += "產量(副產物產量，替代指令：副產物、副產品)\n\n"

        reply += "其他類：\n"
        reply += "作物代碼"
        return reply
    else:
        reply = f"很抱歉，本系統不支援您輸入的指令「{command_text}」，無法為您查詢，請輸入「指令」查詢目前可用的指令。"
        return reply

    try:
        obj = apiview(list_params)
        if hasattr(obj, 'choose_api'):
            obj = obj.choose_api()
        reply = f"搜尋「{command_text}」結果如下\n" + obj.execute_api()
    except CustomError as ce:
        reply = f"無法搜尋「{command_text}」\n" + str(ce)
    except Exception as e:
        traceback_log = TracebackLog.objects.create(app="api_view", message=traceback.format_exc())
        reply = f"搜尋「{command_text}」發生未知錯誤，錯誤編號「{traceback_log.id}」，請通知管理員處理"
        raise CustomError(reply)

    return reply


def file_view_product_code(file_name):
    '''
    處理上傳的代碼excel
    '''
    try:
        ProductCode.objects.all().delete()
        main_count = 0
        labor_count = 0

        wb = load_workbook(file_name)
        ws = wb['主力代碼']
        for row in ws.rows:
            code = row[0]
            name = row[1]
            if code.row <= 1:
                continue
            else:
                obj = ProductCode.objects.create(category='主力', code=code.value, name=name.value)
                main_count += 1

        ws = wb['勞動力代碼']
        for row in ws.rows:
            code = row[0]
            name = row[1]
            if code.row <= 1:
                continue
            else:
                obj = ProductCode.objects.create(category='勞動力', code=code.value, name=name.value)
                labor_count += 1

        return f"上傳主力代碼{main_count}筆，勞動力代碼{labor_count}筆成功！"
    except Exception as e:
        raise


def file_view_crop_produce(file_name):
    '''
    處理上傳的產量產值表
    '''
    CropProduceUnit.objects.all().delete()
    wb = load_workbook(file_name)
    for sheet in wb.sheetnames:
        if "稻" in sheet:
            col_period = "C"
            col_city_district = None
        else:
            col_period = None
            col_city_district = "E"
        col_name = col_city = col_district = col_city_code = col_district_code = None
        col_amount_max = col_amount_min = col_amount_average = col_value_min = col_value_average = None
        amount_unit = value_unit = None

        ws = wb[sheet]
        for cell in ws[1]:
            if cell.value is None:
                continue
            elif "產量" in cell.value and "(公斤)" not in cell.value:
                amount_unit = "(" + cell.value.split("(")[-1]
            elif "產值" in cell.value and "(元)" not in cell.value:
                value_unit = "(" + cell.value.split("(")[-1]

        for cell in ws[2]:
            value = cell.value
            letter = cell.column_letter
            if value is None:
                continue
            elif "農產別" in value or "稻種別" in value:
                col_name = letter
            elif "縣市別" in value:
                col_city = letter
            elif "鄉鎮別" in value and col_period:
                col_city = letter
            elif "鄉鎮別" in value:
                col_district = letter
            elif "縣市代碼" in value:
                col_city_code = letter
            elif "鄉鎮代碼" in value:
                col_district_code = letter
            elif "MAX" in value:
                if col_amount_max is None:
                    col_amount_max = letter
                else:
                    col_value_max = letter
            elif "MIN" in value:
                if col_amount_min is None:
                    col_amount_min = letter
                else:
                    col_value_min = letter
            elif "age" in value:
                if col_amount_average is None:
                    col_amount_average = letter
                else:
                    col_value_average = letter

        for index in range(3, ws.max_row + 1):
            display_name = ws[f"A{index}"].value
            name = ws[f"{col_name}{index}"].value
            city = ws[f"{col_city}{index}"].value
            city_code = ws[f"{col_city_code}{index}"].value
            district_code = ws[f"{col_district_code}{index}"].value
            amount_max = ws[f"{col_amount_max}{index}"].value
            amount_min = ws[f"{col_amount_min}{index}"].value
            amount_average = ws[f"{col_amount_average}{index}"].value
            value_max = ws[f"{col_value_max}{index}"].value
            value_min = ws[f"{col_value_min}{index}"].value
            value_average = ws[f"{col_value_average}{index}"].value

            obj = CropProduceUnit.objects.create(
                category = sheet,
                display_name = display_name,
                name = name,
                city = city,
                city_code = city_code,
                district_code = district_code,
                amount_min = amount_min,
                amount_max = amount_max,
                amount_average = amount_average,
                value_min = value_min,
                value_max = value_max,
                value_average = value_average,
            )
            update = False
            if col_period:
                obj.period = ws[f"{col_period}{index}"].value
                update = True
            if col_district:
                obj.district = ws[f"{col_district}{index}"].value
                update = True
            if col_city_district:
                obj.city_district = ws[f"{col_city_district}{index}"].value
                update = True
            if amount_unit:
                obj.amount_unit = amount_unit
                update = True
            if value_unit:
                obj.value_unit = value_unit
                update = True
            if update:
                obj.save()
    return f"上傳成功！"


def check_database_locked(model_name):
    query_set = DatabaseControl.objects.filter(name=model_name)
    if query_set.count() == 0:
        return

    last_obj = query_set.last()
    wait_time = round(last_obj.expire_time.timestamp() - datetime.datetime.now().timestamp())
    if last_obj.status == True or wait_time < 0:
        return
    else:
        raise CustomError(f"其他使用者更新資料中，請過{wait_time}秒後再嘗試。")


def upload(request):
    user = request.user
    token = request.GET.get('token') or None

    if request.method == "GET":
        if token is not None:
            query_set = AnyToken.objects.filter(token=token)
            if query_set.count() == 0:
                return render(request, 'coa/upload-unauth.html', {'message': '網址無效，請向機器人取得正確的上傳網址'})

            obj_token = query_set.first()
            if obj_token.expire_time < timezone.now():
                return render(request, 'coa/upload-unauth.html', {'message': '網址過期，請重新取得上傳網址'})
        else:
            return render(request, 'coa/upload-unauth.html', {'message': '請向機器人取得上傳網址'})
        user = obj_token.user
        login(request, user)
        return render(request, 'coa/upload.html', locals())

    if request.method == "POST":
        response = ''
        if not isinstance(user, CustomUser):
            response = '網址已過期，請重新取得上傳網址'
            data = {
                'status': 403,
                'error': response,
                'content': response,
            }
            return JsonResponse(data)

        if token is None:
            response = '請向機器人取得上傳網址'
        else:
            query_set = AnyToken.objects.filter(token=token)
            if query_set.count() == 0:
                response = '網址無效，請向機器人取得正確的上傳網址'

            obj_token = query_set.first()
            if obj_token.expire_time < timezone.now():
                response = '網址過期，請重新取得上傳網址'

        if response:
            data = {
                'status': 403,
                'error': response,
                'content': response,
            }
            return JsonResponse(data)

        file = request.FILES.get('file')
        data = file.read()
        filename = file.name
        if "產量" not in filename and "產值" not in filename and "主力" not in filename and "勞動力" not in filename:
            response = f"上傳的檔案名稱「{filename}」不符要求，上傳失敗！"
            data = {
                'status': 500,
                'error': response,
                'content': response,
            }
        else:
            new_filename = filename.split('.')[0] + f"_{datetime.datetime.now().timestamp()}." + filename.split('.')[-1]

            with open(f"{new_filename}", "wb") as f:
                f.write(data)

            try:
                if "產量" in filename:
                    check_database_locked('CropProduceUnit')
                    lock_obj = DatabaseControl.objects.create(user=user, name='CropProduceUnit', expire_time=(timezone.now() + datetime.timedelta(0, 600)))
                    response = f"{filename}" + file_view_crop_produce(new_filename)
                else:
                    check_database_locked('ProductCode')
                    lock_obj = DatabaseControl.objects.create(user=user, name='ProductCode', expire_time=(timezone.now() + datetime.timedelta(0, 600)))
                    response = file_view_product_code(new_filename)
                data = {
                    'status': 200,
                    'success': response,
                    'content': response
                }
                lock_obj.status = True
                lock_obj.finish_time = timezone.now()
                lock_obj.save()
            except CustomError as ce:
                response = str(ce)
                data = {
                    'status': 500,
                    'error': response,
                    'content': response,
                }
            except Exception as e:
                traceback_log = TracebackLog.objects.create(app="upload", message=traceback.format_exc())
                response = f"{filename} 上傳時發生未知錯誤，錯誤編號「{traceback_log.id}」，請通知管理員處理"
                data = {
                    'status': 500,
                    'error': response,
                    'content': response,
                }
            if os.path.exists(new_filename):
                os.remove(new_filename)
        return JsonResponse(data)

    return render(request, 'coa/upload.html', locals())


def change_proxy(command_text, line_user):
    if CustomUser.objects.filter(line_uid=line_user.user_id, is_superuser=True).count() == 0:
        return ''

    list_params = command_text.strip().split(' ')
    if len(list_params) == 1:
        return f"請輸入要更換的代理內容"
    else:
        proxy = list_params[1]

    headless, use_proxy = True, True
    driver = get_driver(headless, use_proxy, proxy)
    driver.set_page_load_timeout(20)
    try:
        driver.get('https://apis.afa.gov.tw/pagepub/AppContentPage.aspx?itemNo=PRI105')
        source = driver.page_source
        if "這個網頁無法正常運作" in source or "重新載入" in source or "詳細資訊" in source:
            return f"{proxy} 無法使用"
        else:
            try:
                obj = CustomSetting.objects.get(name='proxy')
                obj.value = proxy
                obj.save()
            except Exception as e:
                CustomSetting.objects.create(name='proxy', value=proxy)

        driver.close()
        return f"{proxy} 可以使用，已更新"
    except TimeoutException:
        driver.close()
        return f"{proxy} 無法使用"
    except Exception as e:
        driver.close()
        return f"更換代理發生錯誤"


def proxy_parser(request):
    token = request.GET.get('token')
    if token != settings.PROXY_TOKEN:
        return HttpResponse('無法使用此功能')

    uri = request.get_raw_uri()
    data = uri.split('data=')[-1]
    params = urllib.parse.unquote(data).split('&')
    # body = request.body.decode()
    # params = urllib.parse.unquote(body.replace('params=', '')).split('&')
    # return HttpResponse(f"data is {data}\nparams is {params}")
    try:
        obj = CropPriceOriginApiView(params)
        reply = obj.execute_api()
    except Exception as e:
        traceback_log = TracebackLog.objects.create(app="proxy_parser", message=traceback.format_exc())
        reply = str(e)
    return HttpResponse(reply)


def proxy_build(request):
    token = request.GET.get('token')
    api = request.GET.get('api')
    if token != settings.PROXY_TOKEN:
        return HttpResponse('無法使用此功能')

    if api == 'CropPriceOriginBuilder':
        data = CropPriceOriginBuilder().build(use_proxy=True)
    elif api == 'CropProduceTotalBuilder':
        data = CropProduceTotalBuilder().build(use_proxy=True)
    else:
        return HttpResponse('無法使用此功能')
    return JsonResponse(data, safe=False)
